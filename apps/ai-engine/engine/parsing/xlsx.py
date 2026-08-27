from __future__ import annotations

import io

from openpyxl import load_workbook

from engine.parsing.base import ParsedBlock, ParseOutput

MAX_ROWS_PER_SHEET = 2000


def parse_xlsx(data: bytes) -> ParseOutput:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = ParseOutput()
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS_PER_SHEET:
                out.warnings.append(f"sheet {ws.title}: truncated at {MAX_ROWS_PER_SHEET} rows")
                break
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
        if rows:
            flat = "\n".join(" | ".join(r) for r in rows)
            out.blocks.append(
                ParsedBlock("table", flat, section_path=f"Sheet: {ws.title}", table_data=rows)
            )
    return out
